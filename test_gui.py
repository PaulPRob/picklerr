"""Smoke tests for the RR_pickle_picker GUI (runs offscreen, no display).

Uses a temporary players file so the real players.txt roster is untouched.
"""

import importlib
import os

import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PyQt6.QtWidgets import QApplication  # noqa: E402


@pytest.fixture(scope="session")
def app():
    return QApplication.instance() or QApplication([])


@pytest.fixture
def window(app, tmp_path, monkeypatch):
    mod = importlib.import_module("RR_pickle_picker")
    monkeypatch.setattr(mod, "PLAYERS_FILE", tmp_path / "players.txt")
    return mod.PicklePicker()


def add_players(window, n):
    for i in range(1, n + 1):
        window.name_edit.setText(f"Player{i:02d}")
        window._add_player()


def test_add_and_persist_players(window):
    add_players(window, 5)
    assert window.player_list.count() == 5
    saved = importlib.import_module("RR_pickle_picker")
    lines = saved.PLAYERS_FILE.read_text().splitlines()
    assert lines == [f"Player{i:02d}" for i in range(1, 6)]


def test_duplicate_names_not_added(window, monkeypatch):
    from PyQt6.QtWidgets import QMessageBox
    monkeypatch.setattr(QMessageBox, "warning",
                        staticmethod(lambda *a, **k: None))
    add_players(window, 3)
    window.name_edit.setText("player01")  # case-insensitive duplicate
    window._add_player()
    assert window.player_list.count() == 3


def test_counts_label_updates(window):
    add_players(window, 14)
    assert "3 courts" in window.counts_label.text()
    assert "2 byes" in window.counts_label.text()


def test_generate_builds_round_tabs(window):
    add_players(window, 12)
    window.rounds_spin.setValue(4)
    window._generate()
    labels = [window.tabs.tabText(i) for i in range(window.tabs.count())]
    assert labels == ["Round 1", "Round 2", "Round 3", "Round 4", "Summary"]


def test_generate_refuses_too_few_players(window, monkeypatch):
    from PyQt6.QtWidgets import QMessageBox
    warnings = []
    monkeypatch.setattr(QMessageBox, "warning",
                        staticmethod(lambda *a, **k: warnings.append(a)))
    add_players(window, 3)
    window._generate()
    assert warnings
    assert window.tabs.tabText(0) == "Welcome"  # schedule not replaced


def test_export_buttons_enable_after_generate(window):
    buttons = (window.print_btn, window.pdf_btn, window.excel_btn,
               window.html_btn)
    for btn in buttons:
        assert not btn.isEnabled()
    add_players(window, 8)
    window.rounds_spin.setValue(3)
    window._generate()
    for btn in buttons:
        assert btn.isEnabled()


def test_html_export_via_gui(window, tmp_path, monkeypatch):
    from PyQt6.QtWidgets import QFileDialog, QMessageBox
    add_players(window, 10)
    window.rounds_spin.setValue(3)
    window._generate()
    target = tmp_path / "draw"  # .html should be appended automatically
    monkeypatch.setattr(
        QFileDialog, "getSaveFileName",
        staticmethod(lambda *a, **k: (str(target), "")))
    monkeypatch.setattr(QMessageBox, "information",
                        staticmethod(lambda *a, **k: None))
    window._export_html()
    page = (tmp_path / "draw.html").read_text(encoding="utf-8")
    assert page.startswith("<!DOCTYPE html>")
    assert "Round 3" in page


def test_pdf_export_via_gui(window, tmp_path, monkeypatch):
    from PyQt6.QtWidgets import QFileDialog, QMessageBox
    add_players(window, 8)
    window.rounds_spin.setValue(3)
    window._generate()
    target = tmp_path / "out.pdf"
    monkeypatch.setattr(
        QFileDialog, "getSaveFileName",
        staticmethod(lambda *a, **k: (str(target), "PDF files (*.pdf)")))
    monkeypatch.setattr(QMessageBox, "information",
                        staticmethod(lambda *a, **k: None))
    window._export_pdf()
    assert target.read_bytes().startswith(b"%PDF")


def test_excel_export_via_gui(window, tmp_path, monkeypatch):
    from openpyxl import load_workbook
    from PyQt6.QtWidgets import QFileDialog, QMessageBox
    add_players(window, 10)
    window.rounds_spin.setValue(3)
    window._generate()
    target = tmp_path / "out"  # extension should be added automatically
    monkeypatch.setattr(
        QFileDialog, "getSaveFileName",
        staticmethod(lambda *a, **k: (str(target), "")))
    monkeypatch.setattr(QMessageBox, "information",
                        staticmethod(lambda *a, **k: None))
    window._export_excel()
    wb = load_workbook(str(target) + ".xlsx")
    assert wb.sheetnames == ["Schedule", "Summary"]


def test_check_none_and_all(window):
    add_players(window, 6)
    window._check_none()
    assert window._checked_players() == []
    window._check_all()
    assert len(window._checked_players()) == 6
