"""Tests for PDF and Excel export (run offscreen, no display needed)."""

import os

import pytest

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook  # noqa: E402
from PyQt6.QtWidgets import QApplication  # noqa: E402

import export  # noqa: E402
from scheduler import generate_schedule, num_byes  # noqa: E402


@pytest.fixture(scope="session", autouse=True)
def app():
    return QApplication.instance() or QApplication([])


@pytest.fixture(scope="session")
def schedule():
    players = [f"Player{i:02d}" for i in range(1, 15)]  # 14 -> 2 byes/round
    return generate_schedule(players, rounds=5, seed=42)


def test_schedule_html_contains_everything(schedule):
    html = export.schedule_html(schedule)
    for rnd in schedule.rounds:
        assert f"Round {rnd.number}" in html
        for name in rnd.byes:
            assert name in html
    for name in schedule.players:
        assert name in html


def test_schedule_html_escapes_names():
    players = ["A<b>", "B&C", 'D"E', "Fred"]
    sched = generate_schedule(players, rounds=2, seed=1)
    html = export.schedule_html(sched)
    assert "A&lt;b&gt;" in html and "B&amp;C" in html
    assert "A<b>" not in html


def test_export_pdf(schedule, tmp_path):
    path = tmp_path / "sched.pdf"
    export.export_pdf(schedule, str(path))
    data = path.read_bytes()
    assert data.startswith(b"%PDF")
    assert len(data) > 1000


def test_export_excel_roundtrip(schedule, tmp_path):
    path = tmp_path / "sched.xlsx"
    export.export_excel(schedule, str(path))
    wb = load_workbook(str(path))
    assert wb.sheetnames == ["Schedule", "Summary"]

    rows = [[c.value for c in row] for row in wb["Schedule"].iter_rows()]
    flat = "\n".join(str(v) for row in rows for v in row if v is not None)
    # every round header, court header row, and bye appears
    round_headers = [row for row in rows if row and row[0]
                     and str(row[0]).startswith("Round ")]
    assert len(round_headers) == len(schedule.rounds)
    bye_rows = [row for row in rows if row and row[0] == "Byes"]
    assert len(bye_rows) == len(schedule.rounds)  # 14 players -> byes always
    for rnd in schedule.rounds:
        for m in rnd.matches:
            assert f"{m.team1[0]} & {m.team1[1]}" in flat
            assert f"{m.team2[0]} & {m.team2[1]}" in flat
        for b in rnd.byes:
            assert b in flat

    summary = [[c.value for c in row] for row in wb["Summary"].iter_rows()]
    labels = [row[0] for row in summary if row]
    assert "Players" in labels and "Rounds" in labels
    # byes per player listed with correct totals
    total_byes = sum(row[1] for row in summary
                     if row and row[0] in schedule.players)
    assert total_byes == num_byes(len(schedule.players)) * len(schedule.rounds)


def _player_sections(page: str) -> dict[str, str]:
    """Map of player name -> that player's pre-rendered pview section."""
    import re
    sections = {}
    for m in re.finditer(
            r'<section class="pview" id="p-\d+"><h2>(.*?)</h2>(.*?)'
            r'</section>', page, re.S):
        sections[m.group(1)] = m.group(2)
    return sections


def test_export_html_full_draw(schedule, tmp_path):
    path = tmp_path / "draw.html"
    export.export_html(schedule, str(path))
    page = path.read_text(encoding="utf-8")
    assert page.startswith("<!DOCTYPE html>")
    assert '<meta name="viewport"' in page  # mobile layout
    for rnd in schedule.rounds:
        assert f"Round {rnd.number}" in page
        for b in rnd.byes:
            assert b in page
    # one filter chip link per player, plus the Everyone chip
    assert page.count('class="chip"') == len(schedule.players)
    assert page.count('class="chip everyone"') == 1
    # no external resources: fully self-contained for offline use
    assert "http://" not in page and "https://" not in page


def test_export_html_has_no_javascript(schedule, tmp_path):
    # Attachment previews (iOS QuickLook via WhatsApp/Mail) block scripts,
    # so the per-player filter must be pure HTML+CSS.
    path = tmp_path / "draw.html"
    export.export_html(schedule, str(path))
    page = path.read_text(encoding="utf-8")
    assert "<script" not in page.lower()
    assert "onclick" not in page.lower()
    # the CSS :target mechanism that replaces JS must be present
    assert ".pview:target" in page


def test_export_html_player_views_prerendered(schedule, tmp_path):
    path = tmp_path / "draw.html"
    export.export_html(schedule, str(path))
    page = path.read_text(encoding="utf-8")
    sections = _player_sections(page)
    assert sorted(sections) == sorted(schedule.players)
    for rnd in schedule.rounds:
        for m in rnd.matches:
            for name in m.players:
                mine = m.team1 if name in m.team1 else m.team2
                partner = mine[0] if mine[1] == name else mine[1]
                sec = sections[name]
                assert (f"Round {rnd.number}" in sec
                        and f"With <b>{partner}</b>" in sec)
        for name in rnd.byes:
            assert "Bye" in sections[name]
    # every player has exactly one card (game or bye) per round
    for name, sec in sections.items():
        assert sec.count("mine-card") == len(schedule.rounds), name
    # each chip link resolves to an existing section id
    import re
    for target in re.findall(r'href="#(p-\d+)"', page):
        assert f'id="{target}"' in page


def test_export_html_escapes_names(tmp_path):
    players = ["A<b>", "B&C", 'D"E', "Fred"]
    sched = generate_schedule(players, rounds=2, seed=1)
    path = tmp_path / "draw.html"
    export.export_html(sched, str(path))
    page = path.read_text(encoding="utf-8")
    assert "A&lt;b&gt;" in page and "B&amp;C" in page
    assert "A<b>" not in page


def test_export_excel_no_bye_rows_when_divisible(tmp_path):
    sched = generate_schedule([f"P{i}" for i in range(8)], rounds=3, seed=2)
    path = tmp_path / "s.xlsx"
    export.export_excel(sched, str(path))
    rows = [[c.value for c in row]
            for row in load_workbook(str(path))["Schedule"].iter_rows()]
    assert not any(row and row[0] == "Byes" for row in rows)
