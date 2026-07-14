"""Print/export helpers for RR_pickle_picker schedules.

Provides an HTML rendering of a schedule (used for both printing and PDF
export via Qt's rich-text printing), an Excel (.xlsx) export via
openpyxl, and a phone-friendly HTML export.  Kept separate from the GUI
so it can be tested headlessly.
"""

from __future__ import annotations

import datetime
import html

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PyQt6.QtGui import QPageLayout, QTextDocument
from PyQt6.QtPrintSupport import QPrinter

from scheduler import Schedule, compute_stats, num_courts


def schedule_html(schedule: Schedule) -> str:
    """Printable HTML for the whole schedule, one section per round."""
    e = html.escape
    parts = [
        "<h1 align='center'>RR_pickle_picker &mdash; Round Robin</h1>",
        f"<p align='center'>{len(schedule.players)} players &nbsp;|&nbsp; "
        f"{len(schedule.rounds)} rounds &nbsp;|&nbsp; "
        f"{num_courts(len(schedule.players))} courts per round</p>",
    ]
    for rnd in schedule.rounds:
        parts.append(f"<h2>Round {rnd.number}</h2>")
        parts.append(
            "<table border='1' cellspacing='0' cellpadding='6' width='100%'>"
            "<tr><th>Court</th><th>Team 1</th><th>Team 2</th></tr>")
        for i, m in enumerate(rnd.matches, start=1):
            parts.append(
                f"<tr><td align='center'>{i}</td>"
                f"<td align='center'>{e(m.team1[0])} &amp; {e(m.team1[1])}"
                f"</td>"
                f"<td align='center'>{e(m.team2[0])} &amp; {e(m.team2[1])}"
                f"</td></tr>")
        parts.append("</table>")
        if rnd.byes:
            parts.append(
                f"<p><b>Byes:</b> {e(', '.join(rnd.byes))}</p>")
    stats = compute_stats(schedule)
    if any(stats.bye_counts.values()):
        byes = ", ".join(f"{e(p)}: {n}"
                         for p, n in sorted(stats.bye_counts.items()))
        parts.append(f"<h2>Byes per player</h2><p>{byes}</p>")
    return "".join(parts)


def _document(schedule: Schedule) -> QTextDocument:
    doc = QTextDocument()
    doc.setHtml(schedule_html(schedule))
    return doc


def export_pdf(schedule: Schedule, path: str) -> None:
    """Write the schedule to a PDF file (requires a QApplication)."""
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(path)
    printer.setPageOrientation(QPageLayout.Orientation.Portrait)
    _document(schedule).print(printer)


def print_schedule(schedule: Schedule, printer: QPrinter) -> None:
    """Render the schedule to an already-configured printer."""
    _document(schedule).print(printer)


# ---------------------------------------------------------------------------
# Mobile-friendly HTML export (self-contained, per-player filter)
#
# NO JavaScript: restricted viewers such as iOS QuickLook (used by
# WhatsApp/Mail attachment previews) render HTML+CSS but refuse to run
# scripts.  Every player's personal view is pre-rendered, and switching
# views uses plain anchor links with the CSS :target selector, which
# works in those viewers.  If even link taps are blocked, the page
# degrades to showing the full draw.
# ---------------------------------------------------------------------------

_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>RR Pickle Picker — Draw</title>
<style>
:root {
  --bg: #f4f6f4; --card: #ffffff; --ink: #1c2a22; --muted: #5c6f63;
  --accent: #1f6f43; --line: #d9e2db; --bye: #b3541e;
}
@media (prefers-color-scheme: dark) {
  :root {
    --bg: #121a15; --card: #1c2820; --ink: #e8f0ea; --muted: #9db3a5;
    --accent: #3f9d68; --line: #31423a; --bye: #e08a4e;
  }
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: -apple-system, "Segoe UI", Roboto, sans-serif;
  background: var(--bg); color: var(--ink);
  max-width: 640px; margin: 0 auto; padding: 16px 12px 48px;
  font-size: 17px; line-height: 1.45;
}
h1 { font-size: 1.35rem; margin-bottom: 2px; }
.meta { color: var(--muted); font-size: 0.9rem; margin-bottom: 14px; }
.hint { color: var(--muted); font-size: 0.95rem; margin: 10px 0 6px; }
.chips { display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 18px; }
a.chip {
  border: 1px solid var(--line); background: var(--card); color: var(--ink);
  border-radius: 999px; padding: 8px 14px; font-size: 0.95rem;
  text-decoration: none; display: inline-block;
}
a.chip.everyone { background: var(--accent); color: #fff;
  border-color: var(--accent); font-weight: 600; }
.round, .mine-card {
  background: var(--card); border: 1px solid var(--line);
  border-radius: 12px; padding: 12px 14px; margin-bottom: 14px;
}
.round h2 { font-size: 1.1rem; margin-bottom: 8px; }
.court { padding: 8px 0; border-top: 1px solid var(--line); }
.court:first-of-type { border-top: none; }
.court .label { color: var(--muted); font-size: 0.85rem; }
.vs { color: var(--muted); padding: 0 6px; }
.byes { margin-top: 8px; color: var(--bye); font-weight: 600; }
.mine-card { margin-bottom: 10px; }
.mine-card .rnd { font-weight: 700; }
.mine-card .detail { margin-top: 2px; }
.mine-card.bye .detail { color: var(--bye); font-weight: 600; }
.pview { display: none; }
.pview h2 { font-size: 1.2rem; margin-bottom: 2px; color: var(--accent); }
a.back { display: inline-block; color: var(--muted); margin-bottom: 12px; }
.pview:target { display: block; }
.pview:target ~ #all { display: none; }
</style>
</head>
<body id="top">
<h1>&#127955; RR Pickle Picker</h1>
<div class="meta">__META__</div>
<div class="hint">Tap your name to see just your games:</div>
<div class="chips">
  <a class="chip everyone" href="#all">Everyone</a>
__CHIPS__
</div>
__PLAYER_VIEWS__
<section id="all">
__ROUNDS__
</section>
</body>
</html>
"""


def _player_cards(schedule: Schedule, name: str) -> str:
    """Pre-rendered 'my games' cards for one player, one per round."""
    e = html.escape
    cards = []
    for rnd in schedule.rounds:
        match_card = None
        for i, m in enumerate(rnd.matches, start=1):
            if name not in m.players:
                continue
            mine, other = ((m.team1, m.team2) if name in m.team1
                           else (m.team2, m.team1))
            partner = mine[0] if mine[1] == name else mine[1]
            match_card = (
                f'<div class="mine-card">'
                f'<div class="rnd">Round {rnd.number} &middot; '
                f'Court {i}</div>'
                f'<div class="detail">With <b>{e(partner)}</b> '
                f'&nbsp;vs&nbsp; {e(other[0])} &amp; {e(other[1])}'
                f'</div></div>')
            break
        if match_card is None and name in rnd.byes:
            match_card = (
                f'<div class="mine-card bye">'
                f'<div class="rnd">Round {rnd.number}</div>'
                f'<div class="detail">&#9208; Bye — sit this one out'
                f'</div></div>')
        if match_card:
            cards.append(match_card)
    return "".join(cards)


def export_html(schedule: Schedule, path: str) -> None:
    """Write a self-contained, phone-friendly HTML page of the draw.

    Full draw shown round by round; tapping a player's name chip switches
    to a pre-rendered "my games" view (court, partner, opponents, byes
    per round).  Pure HTML+CSS — no JavaScript — so it also works in
    attachment previews (WhatsApp/Mail on iPhone) that block scripts.
    """
    e = html.escape
    meta = (f"{len(schedule.players)} players &middot; "
            f"{len(schedule.rounds)} rounds &middot; "
            f"{num_courts(len(schedule.players))} courts &middot; "
            f"{datetime.date.today():%d %b %Y}")

    ordered = sorted(schedule.players)
    chips = "\n".join(
        f'  <a class="chip" href="#p-{i}">{e(p)}</a>'
        for i, p in enumerate(ordered))

    views = []
    for i, p in enumerate(ordered):
        views.append(
            f'<section class="pview" id="p-{i}">'
            f'<h2>{e(p)}</h2>'
            f'<a class="back" href="#top">&#8678; Back to all names</a>'
            f'{_player_cards(schedule, p)}'
            f'</section>')

    round_blocks = []
    for rnd in schedule.rounds:
        courts = []
        for i, m in enumerate(rnd.matches, start=1):
            courts.append(
                f'<div class="court"><div class="label">Court {i}</div>'
                f'<b>{e(m.team1[0])} &amp; {e(m.team1[1])}</b>'
                f'<span class="vs">vs</span>'
                f'<b>{e(m.team2[0])} &amp; {e(m.team2[1])}</b></div>')
        byes = (f'<div class="byes">&#9208; Byes: {e(", ".join(rnd.byes))}'
                f'</div>' if rnd.byes else "")
        round_blocks.append(
            f'<div class="round"><h2>Round {rnd.number}</h2>'
            f'{"".join(courts)}{byes}</div>')

    page = (_HTML_TEMPLATE
            .replace("__META__", meta)
            .replace("__CHIPS__", chips)
            .replace("__PLAYER_VIEWS__", "\n".join(views))
            .replace("__ROUNDS__", "\n".join(round_blocks)))
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(page)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F6F43")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ROUND_FONT = Font(bold=True, size=12)
CENTER = Alignment(horizontal="center")


def export_excel(schedule: Schedule, path: str) -> None:
    """Write the schedule to an .xlsx workbook.

    Sheet "Schedule": a block per round with court/team rows and byes.
    Sheet "Summary": fairness stats and byes per player.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    row = 1
    ws.cell(row, 1, "RR_pickle_picker — Round Robin").font = Font(
        bold=True, size=14)
    row += 2
    for rnd in schedule.rounds:
        ws.cell(row, 1, f"Round {rnd.number}").font = ROUND_FONT
        row += 1
        for col, title in enumerate(("Court", "Team 1", "Team 2"), start=1):
            c = ws.cell(row, col, title)
            c.fill, c.font, c.alignment = HEADER_FILL, HEADER_FONT, CENTER
        row += 1
        for i, m in enumerate(rnd.matches, start=1):
            ws.cell(row, 1, i).alignment = CENTER
            ws.cell(row, 2, f"{m.team1[0]} & {m.team1[1]}").alignment = CENTER
            ws.cell(row, 3, f"{m.team2[0]} & {m.team2[1]}").alignment = CENTER
            row += 1
        if rnd.byes:
            ws.cell(row, 1, "Byes").font = Font(bold=True)
            ws.cell(row, 2, ", ".join(rnd.byes))
            row += 1
        row += 1  # blank line between rounds
    for col, width in ((1, 10), (2, 32), (3, 32)):
        ws.column_dimensions[get_column_letter(col)].width = width

    stats = compute_stats(schedule)
    summary = wb.create_sheet("Summary")
    rows = [
        ("Players", len(schedule.players)),
        ("Rounds", len(schedule.rounds)),
        ("Courts per round", num_courts(len(schedule.players))),
        ("Pairs partnered more than once", stats.repeat_partnerships),
        ("Most times any pair partnered", stats.max_partner_repeats),
        ("Most times any pair faced each other", stats.max_opponent_repeats),
        ("Same-partner in consecutive rounds",
         stats.consecutive_partner_repeats),
        ("Same-opponent in consecutive rounds",
         stats.consecutive_opponent_repeats),
        (),
        ("Player", "Byes"),
    ]
    for r, values in enumerate(rows, start=1):
        for c, v in enumerate(values, start=1):
            summary.cell(r, c, v)
    summary.cell(len(rows), 1).font = Font(bold=True)
    summary.cell(len(rows), 2).font = Font(bold=True)
    r = len(rows) + 1
    for name in schedule.players:
        summary.cell(r, 1, name)
        summary.cell(r, 2, stats.bye_counts[name])
        r += 1
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 10

    wb.save(path)
